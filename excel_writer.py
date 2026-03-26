# excel_writer.py
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BOLD  = Font(bold=True)


def write_sheet(ws, all_results):
    """Ghi dữ liệu compare vào 1 worksheet"""
    if not all_results:
        ws.cell(row=1, column=1, value="No results (base dir failed or missing Trusses/Presets folder)").font = BOLD
        return

    # Collect ALL sections across all files (preserve order, no duplicates)
    seen = set()
    sections = []
    for _, results in all_results:
        for r in results:
            if r["section"] not in seen:
                seen.add(r["section"])
                sections.append(r["section"])

    # Header row 1
    ws.cell(row=1, column=1, value="File").font = BOLD
    col = 2
    for section in sections:
        ws.cell(row=1, column=col, value=section).font = BOLD
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")
        col += 2
    ws.cell(row=1, column=col, value="Sections with diff").font = BOLD

    # Header row 2
    col = 2
    for _ in sections:
        ws.cell(row=2, column=col,   value="Diff").font = BOLD
        ws.cell(row=2, column=col+1, value="%").font    = BOLD
        col += 2

    # Data rows
    for row_idx, (filename, results) in enumerate(all_results, 3):
        ws.cell(row=row_idx, column=1, value=filename)
        diff_sections = []
        has_any_diff = False
        col = 2

        # Index results by section name for easy lookup
        result_map = {r["section"]: r for r in results}

        for section in sections:
            r = result_map.get(section)
            if r:
                diff_cell = ws.cell(row=row_idx, column=col,   value=r["diff_count"])
                pct_cell  = ws.cell(row=row_idx, column=col+1, value=f"{r['diff_pct']}%")
                if r["diff_count"] > 0:
                    diff_cell.fill = RED
                    pct_cell.fill  = RED
                    diff_sections.append(r["section"])
                    has_any_diff = True
                else:
                    diff_cell.fill = GREEN
                    pct_cell.fill  = GREEN
            else:
                # Section không có trong file này → để trống
                ws.cell(row=row_idx, column=col,   value="-")
                ws.cell(row=row_idx, column=col+1, value="-")
            col += 2

        ws.cell(row=row_idx, column=1).fill = GREEN if not has_any_diff else RED
        ws.cell(row=row_idx, column=col, value=", ".join(diff_sections) if diff_sections else "-")

    # Column widths
    ws.column_dimensions["A"].width = 25
    col = 2
    for _ in sections:
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width   = 10
        ws.column_dimensions[openpyxl.utils.get_column_letter(col+1)].width = 8
        col += 2
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 40


def write_report(base_all_results, output_path):
    """
    base_all_results: dict {base_dir: [(filename, results)]}
    output_path: đường dẫn file Excel xuất ra
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Xóa sheet mặc định

    for i, (base_dir, all_results) in enumerate(base_all_results.items()):
        sheet_name = f"Base Dir {i+1} - {os.path.basename(base_dir)}"
        sheet_name = sheet_name[:31]  # Excel giới hạn 31 ký tự
        ws = wb.create_sheet(title=sheet_name)
        write_sheet(ws, all_results)

    wb.save(output_path)
    print(f"Excel đã xuất: {output_path}")