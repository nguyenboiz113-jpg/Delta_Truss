import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Palette ────────────────────────────────────────────────────────────────────
HDR_DARK    = PatternFill("solid", fgColor="1A2F4A")
HDR_MID     = PatternFill("solid", fgColor="1F3F63")
HDR_ACCENT  = PatternFill("solid", fgColor="2E6DA4")

PASS_FILL   = PatternFill("solid", fgColor="D6F0D6")
FAIL_FILL   = PatternFill("solid", fgColor="FDDEDE")
WARN_FILL   = PatternFill("solid", fgColor="FFF3CD")

ROW_ALT     = PatternFill("solid", fgColor="F7FAFD")
ROW_NORM    = PatternFill("solid", fgColor="FFFFFF")
TOTAL_FILL  = PatternFill("solid", fgColor="E8F0FA")
PROFILE_HDR = PatternFill("solid", fgColor="2D6A9F")

# ── Fonts ──────────────────────────────────────────────────────────────────────
F_HDR        = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
F_HDR_SMALL  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
F_SUB        = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
F_BOLD       = Font(name="Calibri", bold=True, color="1A2F4A", size=10)
F_BOLD_TOTAL = Font(name="Calibri", bold=True, color="1A2F4A", size=11)
F_NORM       = Font(name="Calibri", color="2C2C2C", size=10)
F_MUTED      = Font(name="Calibri", color="888888", size=10, italic=True)

# ── Borders ────────────────────────────────────────────────────────────────────
def _make_border(style="thin", color="C5D0DC"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _make_border_thick_bottom(color="1A2F4A"):
    thick = Side(style="medium", color=color)
    thin  = Side(style="thin",   color="C5D0DC")
    return Border(left=thin, right=thin, top=thin, bottom=thick)

B_NORM  = _make_border()
B_HDR   = _make_border("medium", "0D1E30")
B_THICK = _make_border_thick_bottom()

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── Profile columns ────────────────────────────────────────────────────────────
PROFILE_COLS = [
    ("Truss Label",   "truss_label"),
    ("Plys",          "plys"),
    ("Wind",          "wind"),
    ("Snow",          "snow"),
    ("Status",        "analysis_status"),
    ("Version",       "version"),
    ("Load Template", "load_template"),
]
N_PROFILE = len(PROFILE_COLS)


# ── Helpers ────────────────────────────────────────────────────────────────────
def _set(cell, value=None, font=None, fill=None, border=None, align=None):
    if value is not None: cell.value     = value
    if font:              cell.font      = font
    if fill:              cell.fill      = fill
    if border:            cell.border    = border
    if align:             cell.alignment = align


def _base_dir_names(base_all_results):
    return {
        bd: f"Base Dir {i+1}  ·  {os.path.basename(bd)}"
        for i, bd in enumerate(base_all_results)
    }


def _get_all_sections(base_all_results):
    seen, sections = set(), []
    for _, all_results in base_all_results.items():
        for _, results in all_results:
            for r in results:
                if r["section"] not in seen:
                    seen.add(r["section"])
                    sections.append(r["section"])
    return sections


def _classify_file(results, all_sections):
    responded = {r["section"] for r in results}
    if any(s not in responded for s in all_sections):
        return "not_respond"
    if any(r["diff_count"] > 0 for r in results):
        return "diff"
    return "same"


def _profile_fill(key, value):
    if key == "analysis_status":
        return PASS_FILL if value == "Passed" else FAIL_FILL
    if key in ("wind", "snow"):
        return WARN_FILL if value == "Yes" else ROW_NORM
    return ROW_NORM


# ── Summary Sheet ──────────────────────────────────────────────────────────────
def _write_summary_sheet(ws, base_all_results):
    base_names   = _base_dir_names(base_all_results)
    all_sections = _get_all_sections(base_all_results)

    # Row 1 – title banner
    ws.merge_cells("A1:I1")
    _set(ws["A1"], "COMPARISON REPORT  ─  SUMMARY",
         font=Font(name="Calibri", bold=True, color="FFFFFF", size=14),
         fill=HDR_DARK, border=B_HDR,
         align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 34

    # Row 2 – column headers
    headers   = ["#", "Base Directory", "Total", "✔ Same", "✘ Different", "⚠ Not Respond",
                 "% Same", "% Different", "% Not Respond"]
    widths    = [5, 52, 9, 10, 14, 15, 12, 14, 15]
    hdr_fills = [
        HDR_DARK, HDR_DARK, HDR_DARK,
        PatternFill("solid", fgColor="2E7D32"),
        PatternFill("solid", fgColor="C62828"),
        PatternFill("solid", fgColor="E65100"),
        PatternFill("solid", fgColor="2E7D32"),
        PatternFill("solid", fgColor="C62828"),
        PatternFill("solid", fgColor="E65100"),
    ]
    for col, (h, w, fh) in enumerate(zip(headers, widths, hdr_fills), start=1):
        _set(ws.cell(row=2, column=col), h,
             font=F_HDR_SMALL, fill=fh, border=B_HDR, align=CENTER)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 28

    # Data rows
    summary_data = []
    for base_dir, all_results in base_all_results.items():
        total = len(all_results)
        same = diff = nr = 0
        for _, results in all_results:
            cls = _classify_file(results, all_sections)
            if cls == "same":   same += 1
            elif cls == "diff": diff += 1
            else:               nr   += 1
        summary_data.append((base_names[base_dir], total, same, diff, nr))

    for i, (name, total, same, diff, nr) in enumerate(summary_data, start=1):
        row = i + 2
        alt = ROW_ALT if i % 2 == 0 else ROW_NORM
        p_s  = f"{same/total*100:.1f}%" if total else "0.0%"
        p_d  = f"{diff/total*100:.1f}%" if total else "0.0%"
        p_nr = f"{nr/total*100:.1f}%"   if total else "0.0%"

        row_vals   = [i, name, total, same, diff, nr, p_s, p_d, p_nr]
        row_fills  = [
            alt, alt, alt,
            PASS_FILL,
            FAIL_FILL if diff > 0 else PASS_FILL,
            WARN_FILL if nr   > 0 else PASS_FILL,
            PASS_FILL,
            FAIL_FILL if diff > 0 else PASS_FILL,
            WARN_FILL if nr   > 0 else PASS_FILL,
        ]
        row_fonts  = [
            F_BOLD, F_NORM, F_BOLD,
            Font(name="Calibri", bold=True, color="2E7D32", size=10),
            Font(name="Calibri", bold=True, color="C62828" if diff > 0 else "2E7D32", size=10),
            Font(name="Calibri", bold=True, color="E65100" if nr   > 0 else "2E7D32", size=10),
            Font(name="Calibri", bold=True, color="2E7D32", size=10),
            Font(name="Calibri", bold=True, color="C62828" if diff > 0 else "2E7D32", size=10),
            Font(name="Calibri", bold=True, color="E65100" if nr   > 0 else "2E7D32", size=10),
        ]
        row_aligns = [CENTER, LEFT, CENTER, CENTER, CENTER, CENTER, CENTER, CENTER, CENTER]

        for col, (val, fl, fn, al) in enumerate(zip(row_vals, row_fills, row_fonts, row_aligns), start=1):
            _set(ws.cell(row=row, column=col), val, font=fn, fill=fl, border=B_NORM, align=al)
        ws.row_dimensions[row].height = 20

    # Total row
    tr  = len(summary_data) + 3
    gt  = sum(d[1] for d in summary_data)
    gs  = sum(d[2] for d in summary_data)
    gd  = sum(d[3] for d in summary_data)
    gnr = sum(d[4] for d in summary_data)
    totals = ["", "TOTAL", gt, gs, gd, gnr,
              f"{gs/gt*100:.1f}%"  if gt else "—",
              f"{gd/gt*100:.1f}%"  if gt else "—",
              f"{gnr/gt*100:.1f}%" if gt else "—"]
    for col, val in enumerate(totals, start=1):
        _set(ws.cell(row=tr, column=col), val,
             font=F_BOLD_TOTAL, fill=TOTAL_FILL, border=B_THICK,
             align=LEFT if col == 2 else CENTER)
    ws.row_dimensions[tr].height = 24

    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False


# ── Detail Sheet ───────────────────────────────────────────────────────────────
def write_detail_sheet(ws, base_all_results, base_profiles=None):
    if base_profiles is None:
        base_profiles = {}

    base_names   = _base_dir_names(base_all_results)
    all_sections = _get_all_sections(base_all_results)
    section_start_col = 3 + N_PROFILE
    last_col = section_start_col + len(all_sections) * 2

    # Row 1 – title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    _set(ws.cell(row=1, column=1), "COMPARISON REPORT  ─  DETAIL",
         font=Font(name="Calibri", bold=True, color="FFFFFF", size=13),
         fill=HDR_DARK, border=B_HDR,
         align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 30

    # Row 2 – group headers
    _set(ws.cell(row=2, column=1), "Base Dir",
         font=F_SUB, fill=HDR_MID, border=B_HDR, align=CENTER)
    _set(ws.cell(row=2, column=2), "File",
         font=F_SUB, fill=HDR_MID, border=B_HDR, align=CENTER)

    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=2+N_PROFILE)
    _set(ws.cell(row=2, column=3), "PROFILE",
         font=F_SUB, fill=PROFILE_HDR, border=B_HDR, align=CENTER)

    col = section_start_col
    for section in all_sections:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+1)
        _set(ws.cell(row=2, column=col), section,
             font=F_HDR_SMALL, fill=HDR_ACCENT, border=B_HDR, align=CENTER)
        col += 2

    _set(ws.cell(row=2, column=col), "Sections w/ Diff",
         font=F_SUB, fill=HDR_DARK, border=B_HDR, align=CENTER)
    ws.row_dimensions[2].height = 28

    # Row 3 – sub-headers
    for c_idx in [1, 2]:
        _set(ws.cell(row=3, column=c_idx), fill=HDR_MID, border=B_HDR)

    for i, (label, _) in enumerate(PROFILE_COLS):
        _set(ws.cell(row=3, column=3+i), label,
             font=F_HDR_SMALL, fill=PROFILE_HDR, border=B_HDR, align=CENTER)

    col = section_start_col
    for _ in all_sections:
        for sub, c_col in [("Diff", col), ("%", col+1)]:
            _set(ws.cell(row=3, column=c_col), sub,
                 font=F_SUB, fill=HDR_ACCENT, border=B_HDR, align=CENTER)
        col += 2

    _set(ws.cell(row=3, column=col), fill=HDR_DARK, border=B_HDR)
    ws.row_dimensions[3].height = 20

    # Data rows
    row_idx = 4
    for base_dir, all_results in base_all_results.items():
        base_name = base_names[base_dir]
        profiles  = base_profiles.get(base_dir, {})

        for filename, results in all_results:
            alt = ROW_ALT if row_idx % 2 == 0 else ROW_NORM

            _set(ws.cell(row=row_idx, column=1), base_name,
                 font=F_MUTED, fill=alt, border=B_NORM, align=LEFT)

            file_cell = ws.cell(row=row_idx, column=2)

            profile = profiles.get(filename, {})
            for i, (_, key) in enumerate(PROFILE_COLS):
                value = profile.get(key, "—") if profile else "—"
                pf    = _profile_fill(key, value) if profile else alt
                fn    = F_NORM
                if key == "analysis_status":
                    fn = Font(name="Calibri", bold=True, size=10,
                              color="2E7D32" if value == "Passed" else "C62828")
                _set(ws.cell(row=row_idx, column=3+i), value,
                     font=fn, fill=pf, border=B_NORM, align=CENTER)

            result_map    = {r["section"]: r for r in results}
            diff_sections = []
            has_diff = has_nr = False
            col      = section_start_col

            for section in all_sections:
                r = result_map.get(section)
                if r:
                    is_diff = r["diff_count"] > 0
                    fl = FAIL_FILL if is_diff else PASS_FILL
                    fn = Font(name="Calibri", bold=is_diff, size=10,
                              color="C62828" if is_diff else "2E7D32")
                    _set(ws.cell(row=row_idx, column=col),   r["diff_count"],
                         font=fn, fill=fl, border=B_NORM, align=CENTER)
                    _set(ws.cell(row=row_idx, column=col+1), f"{r['diff_pct']}%",
                         font=fn, fill=fl, border=B_NORM, align=CENTER)
                    if is_diff:
                        diff_sections.append(section)
                        has_diff = True
                else:
                    for c_col in [col, col+1]:
                        _set(ws.cell(row=row_idx, column=c_col), "—",
                             font=F_MUTED, fill=WARN_FILL, border=B_NORM, align=CENTER)
                    has_nr = True
                col += 2

            # File cell styling
            if has_nr:
                file_fill = WARN_FILL
                file_font = Font(name="Calibri", bold=True, color="E65100", size=10)
            elif has_diff:
                file_fill = FAIL_FILL
                file_font = Font(name="Calibri", bold=True, color="C62828", size=10)
            else:
                file_fill = PASS_FILL
                file_font = Font(name="Calibri", bold=True, color="2E7D32", size=10)

            _set(file_cell, filename, font=file_font, fill=file_fill,
                 border=B_NORM, align=LEFT)

            summary_text = (", ".join(diff_sections) if diff_sections
                            else ("— not respond" if has_nr else "—"))
            _set(ws.cell(row=row_idx, column=col), summary_text,
                 font=F_NORM if diff_sections else F_MUTED,
                 fill=FAIL_FILL if diff_sections else (WARN_FILL if has_nr else PASS_FILL),
                 border=B_NORM, align=LEFT)

            ws.row_dimensions[row_idx].height = 18
            row_idx += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 26
    for i, w in enumerate([20, 6, 6, 6, 11, 10, 36]):
        ws.column_dimensions[get_column_letter(3+i)].width = w

    col = section_start_col
    for _ in all_sections:
        ws.column_dimensions[get_column_letter(col)].width   = 8
        ws.column_dimensions[get_column_letter(col+1)].width = 7
        col += 2
    ws.column_dimensions[get_column_letter(col)].width = 42

    ws.freeze_panes = ws.cell(row=4, column=3)
    ws.auto_filter.ref = f"A3:{get_column_letter(section_start_col-1)}3"
    ws.sheet_view.showGridLines = False


# ── Entry point ────────────────────────────────────────────────────────────────
def write_report(base_all_results, output_path, base_profiles=None):
    if base_profiles is None:
        base_profiles = {}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet(title="Summary")
    _write_summary_sheet(ws_summary, base_all_results)

    ws_detail = wb.create_sheet(title="Detail")
    write_detail_sheet(ws_detail, base_all_results, base_profiles)

    wb.save(output_path)
    print(f"Excel exported: {output_path}")